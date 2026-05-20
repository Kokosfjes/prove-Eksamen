from openpyxl import load_workbook
from app import app, db, User

# Leser users.xlsx og lager brukere i databasen.
# Brukernavn = 4 første bokstaver av fornavn (små).
# Passord = brukernavn + "123"

wb = load_workbook("../../users.xlsx")
ark = wb.active

with app.app_context():
    db.create_all()

    for rad in ark.iter_rows(values_only=True):
        tekst = rad[0]
        if not tekst:
            continue

        # Eksempel: "01 Sebastian Bråin Aaen" -> fornavn = "Sebastian"
        deler = tekst.strip().split()
        fornavn = deler[1]
        brukernavn = fornavn[:4].lower()
        passord = brukernavn + "123"

        # Hopp over hvis brukeren allerede finnes
        if User.query.filter_by(username=brukernavn).first():
            continue

        ny = User(username=brukernavn, password=passord)
        db.session.add(ny)
        print(brukernavn, "/", passord)

    db.session.commit()

print("Ferdig.")
