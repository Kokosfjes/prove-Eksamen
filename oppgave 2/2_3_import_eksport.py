# 2.3 Import / eksport
from openpyxl import load_workbook


def lag_brukernavn(navn):
    # Fjerner tall foran navnet (f.eks. "01 Sebastian Aaen" -> "Sebastian Aaen")
    navn = navn.strip().lower()
    deler = navn.split()
    if deler[0].isdigit():
        deler = deler[1:]

    fornavn = deler[0]
    etternavn = deler[-1]
    return fornavn[:3] + etternavn[0]


def lag_epost(navn, domene):
    return lag_brukernavn(navn) + domene


# Les inn navn fra users.xlsx
wb = load_workbook("../../users.xlsx")
ark = wb.active

# Skriv til csv-fil
with open("username_email.csv", "w", encoding="utf-8") as fil:
    for rad in ark.iter_rows(values_only=True):
        if not rad[0]:
            continue
        navn = rad[0]
        brukernavn = lag_brukernavn(navn)
        epost = lag_epost(navn, "@lotus.no")
        fil.write(brukernavn + "; " + epost + "\n")

print("Ferdig. Skrev til username_email.csv")
